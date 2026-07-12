import os
import re
import io
import json
import uuid
import pdfplumber
from flask import Flask, render_template, request, send_file, jsonify, redirect, url_for
from collections import defaultdict
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook, load_workbook
import tempfile
import threading
import database

# ============================================================
# ZONA HORARIA: CIUDAD DE MEXICO
# ============================================================
MX_TZ = ZoneInfo("America/Mexico_City")

def now_mx():
    """Devuelve datetime actual en zona horaria de Ciudad de Mexico"""
    return datetime.now(timezone.utc).astimezone(MX_TZ)

def now_mx_str(fmt="%d/%m/%Y %H:%M:%S"):
    """Devuelve string con fecha/hora de Ciudad de Mexico"""
    return now_mx().strftime(fmt)

# ============================================================
# BASE DE DATOS - Delegada a database.py (Turso/SQLite cloud)
# ============================================================

def init_db():
    """Inicializar base de datos (delegado a database.py)"""
    database.init_db()

def guardar_pedido_db(pedido_id, db):
    """Deprecated - usar database.finalizar_pedido_db() directamente"""
    pass

def cargar_historial_db():
    """Cargar historial (delegado a database.py)"""
    return database.cargar_historial_db()

def get_stats_db():
    """Estadisticas (delegado a database.py)"""
    return database.get_stats_db()
# ============================================================
# GOOGLE DRIVE INTEGRATION
# ============================================================
try:
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload, MediaIoBaseUpload
    from google.oauth2 import service_account
    GOOGLE_DRIVE_ENABLED = True
except ImportError:
    GOOGLE_DRIVE_ENABLED = False

# ID de la carpeta raíz en Google Drive
DRIVE_FOLDER_ID = os.environ.get('GOOGLE_DRIVE_FOLDER_ID', '1HChYtZB51EQkuAlHebmllIju1AUsGI1u')

def get_drive_service():
    """Inicializar servicio de Google Drive"""
    if not GOOGLE_DRIVE_ENABLED:
        return None
    try:
        # Intentar leer desde variable de entorno primero (Render)
        credentials_json = os.environ.get('GOOGLE_DRIVE_CREDENTIALS')

        if credentials_json:
            # Parsear JSON desde variable de entorno
            import json
            credentials_info = json.loads(credentials_json)
            credentials = service_account.Credentials.from_service_account_info(
                credentials_info,
                scopes=['https://www.googleapis.com/auth/drive']
            )
        else:
            # Fallback: leer desde archivo (local development)
            credentials_path = os.path.join(os.path.dirname(__file__), 'credentials.json')
            if not os.path.exists(credentials_path):
                print("No se encontró GOOGLE_DRIVE_CREDENTIALS ni credentials.json")
                return None

            credentials = service_account.Credentials.from_service_account_file(
                credentials_path,
                scopes=['https://www.googleapis.com/auth/drive']
            )

        return build('drive', 'v3', credentials=credentials)
    except Exception as e:
        print(f"Error inicializando Drive: {e}")
        return None

def crear_o_obtener_carpeta(service, nombre, parent_id):
    """Crear carpeta si no existe, o retornar ID si existe"""
    if not service:
        return None

    # Buscar si la carpeta ya existe
    query = f"name='{nombre}' and mimeType='application/vnd.google-apps.folder' and '{parent_id}' in parents and trashed=false"
    results = service.files().list(q=query, spaces='drive', fields='files(id, name)').execute()

    if results['files']:
        return results['files'][0]['id']

    # Crear nueva carpeta
    metadata = {
        'name': nombre,
        'mimeType': 'application/vnd.google-apps.folder',
        'parents': [parent_id]
    }
    folder = service.files().create(body=metadata, fields='id').execute()
    return folder['id']

def subir_a_drive(service, file_path, file_name, folder_id, mime_type='application/pdf'):
    """Subir archivo a Google Drive"""
    if not service or not folder_id:
        return None

    file_metadata = {
        'name': file_name,
        'parents': [folder_id]
    }

    media = MediaFileUpload(file_path, mimetype=mime_type, resumable=True)
    file = service.files().create(body=file_metadata, media_body=media, fields='id, webViewLink').execute()

    return {
        'id': file.get('id'),
        'link': file.get('webViewLink')
    }

def guardar_manifiesto_en_drive(pedido_id, header_data, file_path, file_name, mime_type):
    """Guardar manifiesto en Drive con jerarquía Fecha > Destinatario > Entrega Nro"""
    service = get_drive_service()
    if not service:
        return None

    try:
        # Obtener fecha actual para la carpeta
        fecha = now_mx_str("%Y-%m-%d")
        destinatario = header_data.get('destinatario', 'SIN_DESTINATARIO').replace('/', '-').replace('\\', '-')
        entrega_nro = header_data.get('entrega_nro', 'SIN_ENTREGA').replace('/', '-')

        # Crear jerarquía de carpetas
        fecha_folder_id = crear_o_obtener_carpeta(service, fecha, DRIVE_FOLDER_ID)
        if not fecha_folder_id:
            return None

        dest_folder_id = crear_o_obtener_carpeta(service, destinatario, fecha_folder_id)
        if not dest_folder_id:
            return None

        entrega_folder_id = crear_o_obtener_carpeta(service, entrega_nro, dest_folder_id)
        if not entrega_folder_id:
            return None

        # Subir archivo
        result = subir_a_drive(service, file_path, file_name, entrega_folder_id, mime_type)
        return result

    except Exception as e:
        print(f"Error subiendo a Drive: {e}")
        return None

import qrcode
from PIL import Image as PILImage

app = Flask(__name__)
app.secret_key = "scanner_multi_dispositivo_2026"
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

UPLOAD_FOLDER = tempfile.gettempdir()
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# ============================================================
# INICIALIZAR SQLITE
# ============================================================
init_db()

# ============================================================
# BASE DE DATOS EN MEMORIA (se carga desde SQLite al iniciar)
# ============================================================
PEDIDOS_DB = {}
INVENTARIO_DB = {}

# Cargar pedidos activos desde la base de datos (recovery automatico en restart)
PEDIDOS_DB = database.load_pedidos_activos()

# ============================================================
# LOCK PARA OPERACIONES ATOMICAS (multioperador)
# Evita que 2 operadores se pisen al escanear simultaneamente
# ============================================================
SCAN_LOCK = threading.Lock()

# ============================================================
# ESTADÍSTICAS DE USO (persistencia via SQLite)
# ============================================================
_db_stats = get_stats_db()
USO_STATS = {
    'sesiones_totales': _db_stats['sesiones_totales'],
    'cajas_escaneadas_totales': _db_stats['cajas_escaneadas_totales'],
    'historial': []
}

# ============================================================
# BACKUP AUTOMATICO DIARIO
# ============================================================
def _generar_backup_json():
    """Generar JSON de backup de todos los pedidos activos"""
    backup = {
        'version': 2,
        'fecha_export': now_mx_str("%d/%m/%Y %H:%M:%S"),
        'total_pedidos': len(PEDIDOS_DB),
        'pedidos': {}
    }
    for pid, db in PEDIDOS_DB.items():
        backup['pedidos'][pid] = {
            'pedido_cache': dict(db['pedido_cache']),
            'escaneos_cache': {lote: {
                'cantidad': data['cantidad'],
                'timestamp': list(data['timestamp']),
                'scans': list(data['scans'])
            } for lote, data in db['escaneos_cache'].items()},
            'ultimos_scans': db.get('ultimos_scans', {}),
            'modo_tarima': db.get('modo_tarima', False),
            'tarima_pendiente': db.get('tarima_pendiente', {}),
            'info': db['info']
        }
    return backup

def _guardar_backup_turso():
    """Guardar backup actual en Turso (borra el anterior)"""
    try:
        backup = _generar_backup_json()
        backup_json = json.dumps(backup, ensure_ascii=False, default=str)
        database.save_backup_db(
            fecha=backup['fecha_export'],
            total_pedidos=backup['total_pedidos'],
            backup_json=backup_json
        )
        print(f"[BACKUP] Backup automatico guardado: {backup['total_pedidos']} pedidos")
    except Exception as e:
        print(f"[BACKUP] Error guardando backup: {e}")

def _backup_diario_thread():
    """Hilo que hace backup automatico a las 23:00 cada dia"""
    import time
    while True:
        try:
            ahora = now_mx()
            # Calcular proximas 23:00
            proximo_backup = ahora.replace(hour=23, minute=0, second=0, microsecond=0)
            if ahora >= proximo_backup:
                # Ya paso las 23:00 hoy, programar para manana
                from datetime import timedelta
                proximo_backup = proximo_backup + timedelta(days=1)
            segundos_espera = (proximo_backup - ahora).total_seconds()
            print(f"[BACKUP] Proximo backup automatico: {proximo_backup.strftime('%d/%m/%Y %H:%M')} ({int(segundos_espera)}s)")
            time.sleep(min(segundos_espera, 3600))  # Revisar cada hora max
            # Verificar si ya es hora
            if now_mx().hour >= 23:
                _guardar_backup_turso()
                time.sleep(3600)  # Esperar 1h para no repetir
        except Exception as e:
            print(f"[BACKUP] Error en hilo de backup: {e}")
            time.sleep(3600)

# Guardar backup al iniciar la app (por si se reinicia)
_guardar_backup_turso()

# Iniciar hilo de backup diario
_backup_thread = threading.Thread(target=_backup_diario_thread, daemon=True)
_backup_thread.start()

def parse_cantidad(val_str):
    """Convierte cantidad europea (1.200,00) o americana (1200.00) a float."""
    if not val_str:
        return 0.0
    val_str = val_str.strip()
    if ',' in val_str and '.' in val_str:
        last_comma = val_str.rfind(',')
        last_dot = val_str.rfind('.')
        if last_comma > last_dot:
            val_str = val_str.replace('.', '').replace(',', '.')
        else:
            val_str = val_str.replace(',', '')
    elif ',' in val_str:
        parts = val_str.split(',')
        if len(parts) == 2 and len(parts[1]) <= 2:
            val_str = val_str.replace(',', '.')
        elif len(parts) == 2 and len(parts[0]) > 3:
            val_str = val_str.replace('.', '').replace(',', '.')
    try:
        return float(val_str)
    except:
        return 0.0


# ============================================================
# EXTRAER PICKING PDF CON DATOS DEL HEADER (BLINDADO v4)
# Maneja: cantidades europeas (1.200,00), header multilinea,
#         items tarima/sueltas, descripciones multilinea
# ============================================================
def extraer_picking_pdf(filepath):
    items = []
    header_data = {
        'cliente': '', 'destinatario': '', 'direccion': '',
        'entrega_nro': '', 'fecha_hora_cita': '', 'fecha_hora_impresion': ''
    }

    with pdfplumber.open(filepath) as pdf:
        all_lines = []
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                all_lines.extend(text.split('\n'))

        lines = [l.strip() for l in all_lines if l.strip()]

        # === HEADER ===
        for i, line in enumerate(lines[:20]):
            if 'Cliente:' in line and not header_data['cliente']:
                m = re.search(r'Cliente:\s*(.+?)(?=\s+Fecha/Hora|$)', line)
                if m:
                    header_data['cliente'] = m.group(1).strip()

            if 'Destinatario:' in line and not header_data['destinatario']:
                m = re.search(r'Destinatario:\s*(.+?)(?=\s+Fecha/Hora|$)', line)
                if m:
                    header_data['destinatario'] = m.group(1).strip()

            if ('Dirección' in line or 'Direccion' in line) and not header_data['direccion']:
                m = re.search(r'Direcc[ióo]n\s*(.+)', line)
                if m:
                    header_data['direccion'] = m.group(1).strip()

            if 'Entrega Nro.:' in line and not header_data['entrega_nro']:
                m = re.search(r'Entrega\s*Nro\.:\s*(\S+)', line)
                if m:
                    header_data['entrega_nro'] = m.group(1).strip()

            if 'Fecha/Hora Cita:' in line and not header_data['fecha_hora_cita']:
                m = re.search(r'Fecha/Hora\s*Cita:\s*(.+)', line)
                if m:
                    header_data['fecha_hora_cita'] = m.group(1).strip()

            if ('Fecha/Hora Impresión:' in line or 'Fecha/Hora Impresion:' in line) and not header_data['fecha_hora_impresion']:
                m = re.search(r'Fecha/Hora\s*Impresi[óo]n:\s*(.+)', line)
                if m:
                    header_data['fecha_hora_impresion'] = m.group(1).strip()

        # === ITEMS ===
        item_lines = []
        for i, line in enumerate(lines):
            if re.match(r'^(PT|MP)-\d+', line):
                code_match = re.match(r'^((?:PT|MP)-\d+)(?:\s+(.*))?$', line)
                if code_match:
                    item_lines.append((i, code_match.group(1), code_match.group(2) or ""))

        for idx, (line_num, codigo, desc_resto) in enumerate(item_lines):
            cantidad = None
            lote = None
            calibre = ""
            categoria = ""
            descripcion = desc_resto

            # Descripción multilínea
            next_idx = line_num + 1
            while next_idx < len(lines):
                next_line = lines[next_idx]
                if re.match(r'^(PT|MP)-\d+', next_line):
                    break
                if 'CJ' in next_line:
                    break
                if re.match(r'^Total:', next_line):
                    break
                if re.match(r'^\d+[,.]?\d*\s*$', next_line) and len(next_line) < 15:
                    break
                if next_line in ('de cajas', 'Firma:', 'Encargado Almacén', 'Tarimas',
                                  'Tarimas Sueltas', 'Subtotal resto', 'Picking List',
                                  'Pág.:', 'Código', 'Descripción', 'Subtotal'):
                    break
                descripcion += ' ' + next_line
                next_idx += 1

            max_search = min(8, len(lines) - line_num)
            if idx + 1 < len(item_lines):
                max_search = min(max_search, item_lines[idx + 1][0] - line_num)

            search_text = " ".join(lines[line_num:line_num + max_search])

            pattern_cj = re.search(r'([\d.,]+)\s+CJ\s+(\d+)\s+(\d{10})(?:\s+(.*))?$', search_text)

            if pattern_cj:
                cantidad = parse_cantidad(pattern_cj.group(1))
                lote = pattern_cj.group(3)
                resto_post_lote = pattern_cj.group(4) or ""

                # Extraer descripción limpia: entre código y cantidad
                cantidad_str = pattern_cj.group(1)
                cant_pos = search_text.find(cantidad_str)
                if cant_pos > 0:
                    desc_clean = search_text[len(codigo):cant_pos].strip()
                    if desc_clean and len(desc_clean) < len(descripcion):
                        descripcion = desc_clean

                # Calibre y categoría
                if resto_post_lote:
                    resto_parts = resto_post_lote.strip().split()
                    filtered = []
                    for p in resto_parts:
                        if re.match(r'^\d+[,.]\d+$', p) and len(p) > 3:
                            continue
                        if re.match(r'^\d+$', p):
                            if int(p) <= 200:
                                filtered.append(p)
                        elif p not in ('CJ', 'PT', 'MP'):
                            filtered.append(p)

                    for p in filtered:
                        if re.match(r'^\d+$', p) and not calibre:
                            calibre = p
                        elif re.match(r'^[A-Za-z]\w*$', p) and not categoria:
                            categoria = p

            if cantidad is not None and lote is not None:
                items.append({
                    'codigo': codigo,
                    'descripcion': descripcion.strip() if descripcion else '',
                    'cantidad_pedido': cantidad,
                    'lote': lote,
                    'calibre': calibre,
                    'categoria': categoria
                })

    return items, header_data

# ============================================================
# EXTRAER TRASLADO EXCEL
# Lee archivos Excel de traslado con columnas:
#   Lote | Texto breve material | Total cajas | Embarque | Material | Calibre | Categoria
# Agrupa por lote (suma cajas cuando hay filas duplicadas)
# ============================================================
def extraer_traslado_excel(filepath):
    items = []
    header_data = {
        'cliente': '', 'destinatario': '', 'direccion': '',
        'entrega_nro': '', 'fecha_hora_cita': '', 'fecha_hora_impresion': '',
        'embarque': '', 'tipo': 'TRASLADO_EXCEL'
    }

    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    # Detectar fila de headers (buscar columna "Lote")
    header_row = None
    col_map = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        for col_idx, cell in enumerate(row):
            if cell and isinstance(cell, str):
                cell_lower = cell.strip().lower()
                if cell_lower in ('lote', 'texto breve material', 'total cajas', 'embarque',
                                   'material', 'calibre', 'categoria', 'orden de compra'):
                    header_row = row_idx
                    col_map[cell_lower] = col_idx
        if header_row:
            break

    if not header_row:
        wb.close()
        return items, header_data

    # Leer datos - agrupar por lote
    lotes_agrupados = {}
    embarque_valor = ''
    orden_compra = ''

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or all(v is None or str(v).strip() == '' for v in row[:5]):
            continue

        def get_col(key):
            idx = col_map.get(key)
            return row[idx] if idx is not None and idx < len(row) else None

        lote_raw = get_col('lote')
        if lote_raw is None:
            continue
        lote = str(lote_raw).strip()
        if not lote or lote.lower() == 'nan':
            continue

        descripcion = str(get_col('texto breve material') or '').strip()
        cantidad_raw = get_col('total cajas')
        material = str(get_col('material') or '').strip()
        calibre_raw = get_col('calibre')
        categoria_raw = get_col('categoria')
        embarque_raw = get_col('embarque')
        oc_raw = get_col('orden de compra')

        # Parse cantidad
        try:
            if isinstance(cantidad_raw, (int, float)):
                cantidad = int(cantidad_raw)
            elif isinstance(cantidad_raw, str):
                cantidad = int(parse_cantidad(cantidad_raw))
            else:
                cantidad = 0
        except (ValueError, TypeError):
            cantidad = 0

        if cantidad <= 0:
            continue

        # Guardar embarque (toma el primero no vacio)
        if embarque_raw and not embarque_valor:
            embarque_valor = str(embarque_raw).strip()
        if oc_raw and not orden_compra:
            orden_compra = str(oc_raw).strip()

        calibre = str(calibre_raw).strip() if calibre_raw else ''
        categoria = str(categoria_raw).strip() if categoria_raw else ''

        # Agrupar por lote - sumar cantidades
        if lote not in lotes_agrupados:
            lotes_agrupados[lote] = {
                'codigo': material or 'N/A',
                'descripcion': descripcion,
                'cantidad_pedido': 0,
                'lote': lote,
                'calibre': calibre,
                'categoria': categoria
            }
        lotes_agrupados[lote]['cantidad_pedido'] += cantidad
        # Si tenemos descripcion mejor, actualizar
        if descripcion and not lotes_agrupados[lote]['descripcion']:
            lotes_agrupados[lote]['descripcion'] = descripcion
        if material and (not lotes_agrupados[lote]['codigo'] or lotes_agrupados[lote]['codigo'] == 'N/A'):
            lotes_agrupados[lote]['codigo'] = material

    wb.close()

    items = list(lotes_agrupados.values())
    header_data['embarque'] = embarque_valor
    header_data['entrega_nro'] = orden_compra or embarque_valor
    header_data['cliente'] = f'Traslado - {embarque_valor}' if embarque_valor else 'Traslado Excel'

    return items, header_data

def calcular_estado(pedido, escaneado):
    diff = escaneado - pedido
    if diff == 0:
        return 'OK', 'verde', f'OK'
    elif diff < 0:
        return 'FALTAN', 'rojo', f'FALTAN {abs(int(diff))}'
    else:
        return 'SOBRAN', 'amarillo', f'SOBRAN {int(diff)}'

# ============================================================
# GENERAR QR DE CONEXION
# ============================================================
def generar_qr_conexion(pedido_id, base_url):
    url_operador = f"{base_url}/scan/{pedido_id}"
    qr = qrcode.QRCode(version=1, box_size=10, border=2)
    qr.add_data(url_operador)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#1a5276", back_color="white")

    img_io = io.BytesIO()
    img.save(img_io, 'PNG')
    img_io.seek(0)
    return img_io, url_operador

# ============================================================
# RUTAS: ADMIN
# ============================================================

@app.route("/")
def index():
    return render_template("admin_v3.html")

@app.route("/admin")
def admin():
    return render_template("admin_v3.html")

@app.route("/upload_pdf", methods=["POST"])
def upload_pdf():
    global PEDIDOS_DB, USO_STATS

    USO_STATS['sesiones_totales'] += 1
    try:
        database.increment_sesiones_db()
    except Exception as e:
        print(f"[DB] Error incrementando sesiones: {e}")

    pdf_file = request.files.get("pdf_picking")
    if not pdf_file:
        return jsonify({"error": "Se requiere PDF"}), 400

    pedido_id = str(uuid.uuid4())[:8].upper()
    pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{pedido_id}_{pdf_file.filename}")
    pdf_file.save(pdf_path)

    try:
        picking_items, header_data = extraer_picking_pdf(pdf_path)
    except Exception as e:
        print(f"[PDF] Error extrayendo picking: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error leyendo PDF: {str(e)}"}), 500

    # Detectar si el cliente requiere modo tarima (pedidos masivos)
    cliente = header_data.get('cliente', '').upper()
    modo_tarima = ('COSTCO' in cliente) or ('WAL MART' in cliente) or ('WALMART' in cliente)

    PEDIDOS_DB[pedido_id] = {
        'pedido_cache': {},
        'escaneos_cache': defaultdict(lambda: {'cantidad': 0, 'timestamp': [], 'scans': []}),
        'ultimos_scans': {},
        'modo_tarima': modo_tarima,
        'tarima_pendiente': {},  # {lote: contador_scans} para tracking de 3 scans
        'info': {
            'fecha_creacion': now_mx_str("%d/%m/%Y %H:%M:%S"),
            'nombre_archivo': pdf_file.filename,
            'total_lineas': len(picking_items),
            'inicio_scan': None,
            'fin_scan': None,
            'tiempo_total': None,
            'usuario_operador': None,
            'header_data': header_data
        }
    }

    for item in picking_items:
        lote = item['lote']
        if lote not in PEDIDOS_DB[pedido_id]['pedido_cache']:
            inv_data = INVENTARIO_DB.get(lote, {})
            PEDIDOS_DB[pedido_id]['pedido_cache'][lote] = {
                'codigo': item['codigo'],
                'descripcion': inv_data.get('descripcion', item['descripcion']),
                'calibre': inv_data.get('calibre', item['calibre']),
                'categoria': inv_data.get('categoria', item['categoria']),
                'pedido': 0,
                'escaneado': 0
            }
        PEDIDOS_DB[pedido_id]['pedido_cache'][lote]['pedido'] += item['cantidad_pedido']

    # Persistir pedido en base de datos
    try:
        database.save_pedido_activo(pedido_id, PEDIDOS_DB[pedido_id]['info'], PEDIDOS_DB[pedido_id]['pedido_cache'])
    except Exception as e:
        print(f"[DB] Error guardando pedido activo: {e}")
        import traceback
        traceback.print_exc()

    base_url = request.form.get('base_url', request.host_url.rstrip('/'))
    if 'RENDER_EXTERNAL_URL' in os.environ:
        base_url = os.environ['RENDER_EXTERNAL_URL']

    qr_img, url_operador = generar_qr_conexion(pedido_id, base_url)

    qr_path = os.path.join(app.config['UPLOAD_FOLDER'], f"qr_{pedido_id}.png")
    with open(qr_path, 'wb') as f:
        f.write(qr_img.getvalue())

    resumen = {
        "pedido_id": pedido_id,
        "total_articulos": len(PEDIDOS_DB[pedido_id]['pedido_cache']),
        "total_pedido": sum(v['pedido'] for v in PEDIDOS_DB[pedido_id]['pedido_cache'].values()),
        "total_escaneado": 0,
        "ok": 0,
        "faltantes": len(PEDIDOS_DB[pedido_id]['pedido_cache']),
        "sobrantes": 0,
        "header_data": header_data
    }

    return jsonify({
        "success": True,
        "pedido_id": pedido_id,
        "resumen": resumen,
        "url_operador": url_operador,
        "qr_url": f"/qr_image/{pedido_id}",
        "modo_tarima": modo_tarima
    })

@app.route("/upload_excel", methods=["POST"])
def upload_excel():
    global PEDIDOS_DB, USO_STATS

    USO_STATS['sesiones_totales'] += 1
    try:
        database.increment_sesiones_db()
    except Exception as e:
        print(f"[DB] Error incrementando sesiones: {e}")

    excel_file = request.files.get("excel_traslado")
    if not excel_file:
        return jsonify({"error": "Se requiere archivo Excel"}), 400

    pedido_id = str(uuid.uuid4())[:8].upper()
    excel_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{pedido_id}_{excel_file.filename}")
    excel_file.save(excel_path)

    try:
        picking_items, header_data = extraer_traslado_excel(excel_path)
    except Exception as e:
        return jsonify({"error": f"Error leyendo Excel: {str(e)}"}), 400

    if not picking_items:
        return jsonify({"error": "No se encontraron items validos en el Excel. Verifica que tenga las columnas: Lote, Texto breve material, Total cajas"}), 400

    PEDIDOS_DB[pedido_id] = {
        'pedido_cache': {},
        'escaneos_cache': defaultdict(lambda: {'cantidad': 0, 'timestamp': [], 'scans': []}),
        'ultimos_scans': {},
        'info': {
            'fecha_creacion': now_mx_str("%d/%m/%Y %H:%M:%S"),
            'nombre_archivo': excel_file.filename,
            'total_lineas': len(picking_items),
            'inicio_scan': None,
            'fin_scan': None,
            'tiempo_total': None,
            'usuario_operador': None,
            'header_data': header_data
        }
    }

    for item in picking_items:
        lote = item['lote']
        if lote not in PEDIDOS_DB[pedido_id]['pedido_cache']:
            inv_data = INVENTARIO_DB.get(lote, {})
            PEDIDOS_DB[pedido_id]['pedido_cache'][lote] = {
                'codigo': item['codigo'],
                'descripcion': inv_data.get('descripcion', item['descripcion']),
                'calibre': inv_data.get('calibre', item['calibre']),
                'categoria': inv_data.get('categoria', item['categoria']),
                'pedido': 0,
                'escaneado': 0
            }
        PEDIDOS_DB[pedido_id]['pedido_cache'][lote]['pedido'] += item['cantidad_pedido']

    # Persistir pedido en base de datos
    try:
        database.save_pedido_activo(pedido_id, PEDIDOS_DB[pedido_id]['info'], PEDIDOS_DB[pedido_id]['pedido_cache'])
    except Exception as e:
        print(f"[DB] Error guardando pedido activo: {e}")
        import traceback
        traceback.print_exc()

    base_url = request.form.get('base_url', request.host_url.rstrip('/'))
    if 'RENDER_EXTERNAL_URL' in os.environ:
        base_url = os.environ['RENDER_EXTERNAL_URL']

    qr_img, url_operador = generar_qr_conexion(pedido_id, base_url)

    qr_path = os.path.join(app.config['UPLOAD_FOLDER'], f"qr_{pedido_id}.png")
    with open(qr_path, 'wb') as f:
        f.write(qr_img.getvalue())

    resumen = {
        "pedido_id": pedido_id,
        "total_articulos": len(PEDIDOS_DB[pedido_id]['pedido_cache']),
        "total_pedido": sum(v['pedido'] for v in PEDIDOS_DB[pedido_id]['pedido_cache'].values()),
        "total_escaneado": 0,
        "ok": 0,
        "faltantes": len(PEDIDOS_DB[pedido_id]['pedido_cache']),
        "sobrantes": 0,
        "header_data": header_data
    }

    return jsonify({
        "success": True,
        "pedido_id": pedido_id,
        "resumen": resumen,
        "url_operador": url_operador,
        "qr_url": f"/qr_image/{pedido_id}",
        "modo_tarima": False
    })

@app.route("/qr_image/<pedido_id>")
def qr_image(pedido_id):
    qr_path = os.path.join(app.config['UPLOAD_FOLDER'], f"qr_{pedido_id}.png")
    if os.path.exists(qr_path):
        return send_file(qr_path, mimetype='image/png')
    return "QR no encontrado", 404

@app.route("/get_status/<pedido_id>")
def get_status(pedido_id):
    if pedido_id not in PEDIDOS_DB:
        return jsonify({"error": "Pedido no encontrado"}), 404

    db = PEDIDOS_DB[pedido_id]
    pedido_cache = db['pedido_cache']
    escaneos_cache = db['escaneos_cache']

    resultados = []
    ok = faltantes = sobrantes = 0

    for lote, data in pedido_cache.items():
        estado, color, texto = calcular_estado(data['pedido'], data['escaneado'])
        if color == 'verde': ok += 1
        elif color == 'rojo': faltantes += 1
        elif color == 'amarillo': sobrantes += 1

        scans = escaneos_cache[lote]['scans']
        ultimo = scans[-1]['hora'] if scans else '-'

        resultados.append({
            'codigo': data['codigo'],
            'descripcion': data['descripcion'],
            'lote': lote,
            'calibre': data['calibre'],
            'categoria': data['categoria'],
            'pedido': data['pedido'],
            'escaneado': data['escaneado'],
            'diferencia': data['escaneado'] - data['pedido'],
            'estado': texto,
            'color': color,
            'ultimo_scan': ultimo,
            'num_scans': len(scans)
        })

    sobrantes_pdf = 0
    for lote, data in escaneos_cache.items():
        if lote not in pedido_cache:
            sobrantes_pdf += 1
            inv_data = INVENTARIO_DB.get(lote, {})
            resultados.append({
                'codigo': inv_data.get('codigo', 'NO EN PEDIDO'),
                'descripcion': inv_data.get('descripcion', 'Producto no listado'),
                'lote': lote,
                'calibre': inv_data.get('calibre', '-'),
                'categoria': inv_data.get('categoria', '-'),
                'pedido': 0,
                'escaneado': data['cantidad'],
                'diferencia': data['cantidad'],
                'estado': 'SOBRANTE INESPERADO',
                'color': 'amarillo',
                'ultimo_scan': data['timestamp'][-1] if data['timestamp'] else '-',
                'num_scans': len(data['scans'])
            })

    def sort_key(r):
        if r['color'] == 'rojo': return 0
        if r['color'] == 'amarillo': return 1
        return 2
    resultados.sort(key=sort_key)

    # ============================================================
    # METRICAS SEGUN LOGICA OPERATIVA:
    #
    # ARTICULOS (del pedido):
    #   Articulos OK = items del pedido con al menos 1 caja escaneada
    #   Articulos Faltan = items del pedido sin ninguna caja escaneada
    #   Formula: Articulos OK + Articulos Faltan = Total Articulos
    #   (Los articulos fuera de pedido NO se suman al total)
    #
    # CAJAS:
    #   Cajas OK = cajas escaneadas que estan en el pedido
    #   Cajas Fuera Pedido = cajas escaneadas que NO estan en el pedido
    #   Cajas Escaneadas = total de todas las cajas escaneadas
    #   Formula: Cajas OK + Cajas Fuera = Cajas Escaneadas
    # ============================================================

    # Contar articulos del pedido (sin fuera de pedido)
    articulos_del_pedido = [r for r in resultados if r['estado'] != 'SOBRANTE INESPERADO']
    articulos_ok = sum(1 for r in articulos_del_pedido if r['escaneado'] > 0)
    articulos_faltan = sum(1 for r in articulos_del_pedido if r['escaneado'] == 0)
    total_articulos_pedido = len(articulos_del_pedido)

    # Cajas
    cajas_ok = sum(r['escaneado'] for r in articulos_del_pedido)
    cajas_escaneadas = sum(r['escaneado'] for r in resultados)
    cajas_fuera_pedido = sum(r['escaneado'] for r in resultados if r['estado'] == 'SOBRANTE INESPERADO')
    articulos_fuera_pedido = sum(1 for r in resultados if r['estado'] == 'SOBRANTE INESPERADO')

    # Nueva metrica: Cajas Faltantes por escanear
    cajas_faltantes_por_escanear = max(0, sum(r['pedido'] for r in articulos_del_pedido) - cajas_ok)

    # ============================================================
    # METRICAS POR OPERADOR (multioperador)
    # Contar cuantas cajas escaneo cada usuario
    # ============================================================
    scans_por_operador = {}
    usuarios_unicos = set()
    for lote_data in escaneos_cache.values():
        for scan in lote_data.get('scans', []):
            u = scan.get('usuario', 'SIN_USUARIO')
            if u:
                usuarios_unicos.add(u)
                scans_por_operador[u] = scans_por_operador.get(u, 0) + scan.get('cantidad', 1)

    # Lista de operadores activos (nueva forma, multioperador)
    usuarios_lista = db['info'].get('usuarios_operadores', [])
    if not usuarios_lista and db['info'].get('usuario_operador'):
        usuarios_lista = [db['info']['usuario_operador']]

    resumen = {
        "pedido_id": pedido_id,
        "total_articulos": total_articulos_pedido,
        "total_pedido": sum(r['pedido'] for r in articulos_del_pedido),
        "total_escaneado": cajas_escaneadas,
        "cajas_ok": cajas_ok,
        "cajas_escaneadas": cajas_escaneadas,
        "cajas_fuera_pedido": cajas_fuera_pedido,
        "cajas_faltantes_por_escanear": cajas_faltantes_por_escanear,
        "articulos_fuera_pedido": articulos_fuera_pedido,
        "ok": articulos_ok,
        "faltantes": articulos_faltan,
        "sobrantes": 0,
        "fecha": now_mx_str("%d/%m/%Y %H:%M:%S"),
        "inicio_scan": db['info']['inicio_scan'],
        "fin_scan": db['info']['fin_scan'],
        "tiempo_total": db['info']['tiempo_total'],
        "usuario_operador": db['info']['usuario_operador'],
        "usuarios_operadores": usuarios_lista,
        "scans_por_operador": scans_por_operador,
        "num_operadores": len(usuarios_lista),
        "header_data": db['info']['header_data']
    }

    return jsonify({
        "resumen": resumen,
        "resultados": resultados,
        "info": db['info'],
        "modo_tarima": db.get('modo_tarima', False)
    })

# ============================================================
# RUTAS: HISTORIAL Y MONITOREO
# ============================================================

@app.route("/historial")
def historial_pedidos():
    """Obtener historial completo de pedidos (activos + SQLite)"""
    historial_db = cargar_historial_db()

    # Combinar con pedidos activos que aun no estan en DB
    activos_no_db = []
    for pid, db_data in PEDIDOS_DB.items():
        ya_en_db = any(h['pedido_id'] == pid for h in historial_db)
        if not ya_en_db:
            info = db_data['info']
            header = info.get('header_data', {})
            pedido_cache = db_data['pedido_cache']
            total_escaneado = sum(v['escaneado'] for v in pedido_cache.values())
            activos_no_db.append({
                'pedido_id': pid,
                'nombre_archivo': info.get('nombre_archivo', ''),
                'header_data': header,
                'fecha_creacion': info.get('fecha_creacion', ''),
                'fecha_finalizacion': info.get('fin_scan', ''),
                'usuario_operador': info.get('usuario_operador', ''),
                'tiempo_total': info.get('tiempo_total', ''),
                'embarque': header.get('embarque', ''),
                'en_proceso': info.get('fin_scan') is None,
                'total_escaneado': total_escaneado
            })

    return jsonify({
        "success": True,
        "historial_db": historial_db,
        "activos": activos_no_db,
        "stats": get_stats_db()
    })

# ============================================================
# RUTAS: INVENTARIO
# ============================================================

@app.route("/upload_inventario", methods=["POST"])
def upload_inventario():
    global INVENTARIO_DB

    file = request.files.get("inventario")
    if not file:
        return jsonify({"error": "Se requiere archivo"}), 400

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(filepath)

    try:
        import csv
        count = 0

        if file.filename.endswith('.csv'):
            with open(filepath, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    lote = str(row.get('Lote', '')).strip()
                    if not lote:
                        for key in row.keys():
                            if 'lote' in key.lower():
                                lote = str(row[key]).strip()
                                break

                    if lote and lote != 'nan':
                        INVENTARIO_DB[lote] = {
                            'codigo': str(row.get('Codigo', row.get('codigo', ''))),
                            'descripcion': str(row.get('Descripcion', row.get('descripcion', ''))),
                            'calibre': str(row.get('Calibre', row.get('calibre', ''))),
                            'categoria': str(row.get('Categoria', row.get('categoria', '')))
                        }
                        count += 1
        else:
            wb = load_workbook(filepath)
            ws = wb.active

            headers = [cell.value for cell in ws[1]]
            lote_idx = None
            for i, h in enumerate(headers):
                if h and 'lote' in str(h).lower():
                    lote_idx = i
                    break

            if lote_idx is None:
                return jsonify({"error": "No se encontró columna 'Lote'"}), 400

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[lote_idx]:
                    lote = str(row[lote_idx]).strip()
                    if lote and lote != 'nan':
                        INVENTARIO_DB[lote] = {
                            'codigo': str(row[headers.index('Codigo')] if 'Codigo' in headers else ''),
                            'descripcion': str(row[headers.index('Descripcion')] if 'Descripcion' in headers else ''),
                            'calibre': str(row[headers.index('Calibre')] if 'Calibre' in headers else ''),
                            'categoria': str(row[headers.index('Categoria')] if 'Categoria' in headers else '')
                        }
                        count += 1

        return jsonify({
            "success": True,
            "total_registros": count,
            "mensaje": f"Inventario cargado: {count} lotes"
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ============================================================
# RUTAS: TIMESTAMPS (INICIO/FIN) - AHORA DESDE OPERADOR
# ============================================================

@app.route("/iniciar_scan/<pedido_id>", methods=["POST"])
def iniciar_scan(pedido_id):
    if pedido_id not in PEDIDOS_DB:
        return jsonify({"error": "Pedido no encontrado"}), 404

    data = request.get_json() or {}
    usuario = data.get('usuario', '')

    PEDIDOS_DB[pedido_id]['info']['inicio_scan'] = now_mx_str("%d/%m/%Y %H:%M:%S")
    PEDIDOS_DB[pedido_id]['info']['fin_scan'] = None
    PEDIDOS_DB[pedido_id]['info']['tiempo_total'] = None
    PEDIDOS_DB[pedido_id]['info']['usuario_operador'] = usuario

    # Persistir cambio en base de datos
    try:
        database.update_pedido_info(pedido_id, PEDIDOS_DB[pedido_id]['info'])
    except Exception as e:
        print(f"[DB] Error actualizando info: {e}")

    return jsonify({
        "success": True,
        "inicio": PEDIDOS_DB[pedido_id]['info']['inicio_scan']
    })

@app.route("/finalizar_scan/<pedido_id>", methods=["POST"])
def finalizar_scan(pedido_id):
    if pedido_id not in PEDIDOS_DB:
        return jsonify({"error": "Pedido no encontrado"}), 404

    fin = now_mx()
    PEDIDOS_DB[pedido_id]['info']['fin_scan'] = fin.strftime("%d/%m/%Y %H:%M:%S")

    inicio_str = PEDIDOS_DB[pedido_id]['info']['inicio_scan']
    if inicio_str:
        inicio = datetime.strptime(inicio_str, "%d/%m/%Y %H:%M:%S")
        diff = fin - inicio
        horas = int(diff.total_seconds() // 3600)
        minutos = int((diff.total_seconds() % 3600) // 60)
        segundos = int(diff.total_seconds() % 60)
        PEDIDOS_DB[pedido_id]['info']['tiempo_total'] = f"{horas:02d}:{minutos:02d}:{segundos:02d}"

    # Guardar en estadísticas
    db = PEDIDOS_DB[pedido_id]
    total_escaneado = sum(v['escaneado'] for v in db['pedido_cache'].values())
    USO_STATS['cajas_escaneadas_totales'] += total_escaneado
    try:
        database.update_cajas_escaneadas_db(total_escaneado)
    except Exception as e:
        print(f"[DB] Error actualizando cajas: {e}")
    USO_STATS['historial'].append({
        'pedido_id': pedido_id,
        'cliente': db['info']['header_data'].get('cliente', ''),
        'destinatario': db['info']['header_data'].get('destinatario', ''),
        'entrega_nro': db['info']['header_data'].get('entrega_nro', ''),
        'usuario': db['info']['usuario_operador'],
        'inicio': db['info']['inicio_scan'],
        'fin': db['info']['fin_scan'],
        'tiempo_total': db['info']['tiempo_total'],
        'cajas_escaneadas': total_escaneado
    })

    # Persistir finalizacion en base de datos
    try:
        database.update_pedido_info(pedido_id, PEDIDOS_DB[pedido_id]['info'])
    except Exception as e:
        print(f"[DB] Error actualizando finalizacion: {e}")

    return jsonify({
        "success": True,
        "fin": PEDIDOS_DB[pedido_id]['info']['fin_scan'],
        "tiempo_total": PEDIDOS_DB[pedido_id]['info']['tiempo_total']
    })

# ============================================================
# RUTAS: ESTADÍSTICAS Y MONITOREO
# ============================================================

@app.route("/get_stats")
def get_stats():
    return jsonify({
        "sesiones_totales": USO_STATS['sesiones_totales'],
        "cajas_escaneadas_totales": USO_STATS['cajas_escaneadas_totales'],
        "pedidos_activos": sum(1 for p in PEDIDOS_DB.values() if p['info']['fin_scan'] is None),
        "pedidos_completados": sum(1 for p in PEDIDOS_DB.values() if p['info']['fin_scan'] is not None)
    })

@app.route("/get_monitoreo")
def get_monitoreo():
    pedidos = []
    pedidos_vistos = set()

    # 1. Pedidos activos en memoria
    for pid, db in PEDIDOS_DB.items():
        header = db['info']['header_data']
        total_pedido = sum(v['pedido'] for v in db['pedido_cache'].values())
        total_escaneado = sum(v['escaneado'] for v in db['pedido_cache'].values())
        faltantes = sum(1 for v in db['pedido_cache'].values() if v['escaneado'] < v['pedido'])
        sobrantes = sum(1 for lote, data in db['escaneos_cache'].items() if lote not in db['pedido_cache'])

        status = 'CERRADO' if db['info']['fin_scan'] else 'ABIERTO'
        pedidos_vistos.add(pid)

        pedidos.append({
            'pedido_id': pid,
            'destinatario': header.get('destinatario', ''),
            'entrega_nro': header.get('entrega_nro', ''),
            'cliente': header.get('cliente', ''),
            'embarque': header.get('embarque', ''),
            'cajas_totales': total_pedido,
            'cajas_escaneadas': total_escaneado,
            'cajas_faltantes': total_pedido - total_escaneado if total_pedido > total_escaneado else 0,
            'cajas_sobrantes': sobrantes,
            'inicio': db['info']['inicio_scan'] or '-',
            'fin': db['info']['fin_scan'] or '-',
            'tiempo_total': db['info']['tiempo_total'] or '-',
            'status': status,
            'usuario': db['info']['usuario_operador'] or '-'
        })

    # 2. Pedidos históricos de SQLite (los que no están activos en memoria)
    try:
        historial_db = cargar_historial_db()
        for h in historial_db:
            if h['pedido_id'] in pedidos_vistos:
                continue
            header = h['header_data']
            resumen = h['resumen']
            pedidos.append({
                'pedido_id': h['pedido_id'],
                'destinatario': header.get('destinatario', ''),
                'entrega_nro': header.get('entrega_nro', ''),
                'cliente': header.get('cliente', ''),
                'embarque': header.get('embarque', ''),
                'cajas_totales': resumen.get('total_pedido', 0),
                'cajas_escaneadas': resumen.get('total_escaneado', 0),
                'cajas_faltantes': resumen.get('total_pedido', 0) - resumen.get('cajas_ok', 0),
                'cajas_sobrantes': resumen.get('cajas_fuera_pedido', 0),
                'inicio': h.get('fecha_creacion', '-'),
                'fin': h.get('fecha_finalizacion', '-'),
                'tiempo_total': h.get('tiempo_total', '-'),
                'status': 'CERRADO',
                'usuario': h.get('usuario_operador', '-')
            })
    except Exception as e:
        print(f"[DB] Error cargando monitoreo historico: {e}")

    return jsonify({"pedidos": pedidos})

@app.route("/download_uso_acumulado")
def download_uso_acumulado():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Uso Acumulado'

    headers = ['Pedido ID', 'Inicio', 'Fin', 'Usuario', 'Cliente', 'Destinatario',
               'Embarque', 'Nro Entrega', 'Cajas Pedido', 'Cajas Escaneadas',
               'Cajas OK', 'Cajas Fuera', 'Tiempo Total']
    ws.append(headers)

    # Cargar desde SQLite (persistente) + memoria
    total_cajas = 0
    total_sesiones = 0

    # Primero: pedidos finalizados en memoria
    for pid, db in PEDIDOS_DB.items():
        info = db['info']
        if info['fin_scan']:
            header = info.get('header_data', {})
            pedido_cache = db['pedido_cache']
            cajas_ped = sum(v['pedido'] for v in pedido_cache.values())
            cajas_esc = sum(v['escaneado'] for v in pedido_cache.values())
            cajas_ok = sum(v['escaneado'] for v in pedido_cache.values() if v['escaneado'] > 0)
            cajas_fuera = sum(1 for l, d in db['escaneos_cache'].items() if l not in pedido_cache)
            ws.append([
                pid, info.get('inicio_scan', ''), info['fin_scan'],
                info.get('usuario_operador', ''), header.get('cliente', ''),
                header.get('destinatario', ''), header.get('embarque', ''),
                header.get('entrega_nro', ''), cajas_ped, cajas_esc,
                cajas_ok, cajas_fuera, info.get('tiempo_total', '')
            ])
            total_cajas += cajas_esc
            total_sesiones += 1

    # Segundo: historial de SQLite
    historial_db = cargar_historial_db()
    for h in historial_db:
        header = h['header_data']
        resumen = h['resumen']
        ws.append([
            h['pedido_id'], h.get('fecha_creacion', ''), h.get('fecha_finalizacion', ''),
            h.get('usuario_operador', ''), header.get('cliente', ''),
            header.get('destinatario', ''), header.get('embarque', ''),
            header.get('entrega_nro', ''), resumen.get('total_pedido', 0),
            resumen.get('total_escaneado', 0), resumen.get('cajas_ok', 0),
            resumen.get('cajas_fuera_pedido', 0), h.get('tiempo_total', '')
        ])
        total_cajas += resumen.get('total_escaneado', 0)
        total_sesiones += 1

    # Totales
    ws.append([])
    ws.append(['TOTALES', '', '', '', '', '', '', '', '', total_cajas, '', '', ''])
    ws.append(['TOTAL SESIONES', total_sesiones, '', '', '', '', '', '', '', '', '', '', ''])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name="Uso_Acumulado.xlsx", as_attachment=True)

@app.route("/download_detalle_sesiones")
def download_detalle_sesiones():
    """Exportar Excel con detalle por sesion: item, cajas pedido, cajas escaneadas, diferencia"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Detalle por Sesion'

    headers = ['Sesion ID', 'Cliente', 'Destinatario', 'Embarque', 'Entrega Nro',
               'Fecha Inicio', 'Fecha Fin', 'Usuario', 'Status',
               'Codigo', 'Descripcion', 'Lote', 'Calibre', 'Categoria',
               'Cajas Packing List', 'Cajas Escaneadas', 'Diferencia', 'Estado']
    ws.append(headers)

    # 1. Pedidos activos en memoria (incluye abiertos y cerrados)
    for pid, db in PEDIDOS_DB.items():
        info = db['info']
        header = info.get('header_data', {})
        status = 'CERRADO' if info['fin_scan'] else 'ABIERTO'
        pedido_cache = db['pedido_cache']
        escaneos_cache = db['escaneos_cache']

        for lote, data in pedido_cache.items():
            estado, color, texto = calcular_estado(data['pedido'], data['escaneado'])
            ws.append([
                pid,
                header.get('cliente', ''),
                header.get('destinatario', ''),
                header.get('embarque', ''),
                header.get('entrega_nro', ''),
                info.get('inicio_scan', '') or '',
                info.get('fin_scan', '') or '',
                info.get('usuario_operador', '') or '',
                status,
                data['codigo'],
                data['descripcion'],
                lote,
                data['calibre'],
                data['categoria'],
                data['pedido'],
                data['escaneado'],
                data['escaneado'] - data['pedido'],
                texto
            ])

        # Items fuera de pedido (sobrantes)
        for lote, data in escaneos_cache.items():
            if lote not in pedido_cache:
                inv_data = INVENTARIO_DB.get(lote, {})
                ws.append([
                    pid,
                    header.get('cliente', ''),
                    header.get('destinatario', ''),
                    header.get('embarque', ''),
                    header.get('entrega_nro', ''),
                    info.get('inicio_scan', '') or '',
                    info.get('fin_scan', '') or '',
                    info.get('usuario_operador', '') or '',
                    status,
                    inv_data.get('codigo', 'NO EN PEDIDO'),
                    inv_data.get('descripcion', 'Producto no listado'),
                    lote,
                    inv_data.get('calibre', '-'),
                    inv_data.get('categoria', '-'),
                    0,
                    data['cantidad'],
                    data['cantidad'],
                    'SOBRANTE INESPERADO'
                ])

    # 2. Pedidos historicos en base de datos
    try:
        historial_db = cargar_historial_db()
        for h in historial_db:
            header = h['header_data']
            resultados = h.get('resultados', [])
            for r in resultados:
                ws.append([
                    h['pedido_id'],
                    header.get('cliente', ''),
                    header.get('destinatario', ''),
                    header.get('embarque', ''),
                    header.get('entrega_nro', ''),
                    h.get('fecha_creacion', '') or '',
                    h.get('fecha_finalizacion', '') or '',
                    h.get('usuario_operador', '') or '',
                    'CERRADO',
                    r.get('codigo', ''),
                    r.get('descripcion', ''),
                    r.get('lote', ''),
                    r.get('calibre', ''),
                    r.get('categoria', ''),
                    r.get('pedido', 0),
                    r.get('escaneado', 0),
                    r.get('diferencia', 0),
                    r.get('estado', '')
                ])
    except Exception as e:
        print(f"[DB] Error cargando historial para detalle: {e}")

    # Auto-ajustar columnas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 30)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name="Detalle_Sesiones.xlsx", as_attachment=True)

# ============================================================
# BACKUP / RESTORE - Copia de seguridad de sesiones activas
# ============================================================

@app.route("/download_backup")
def download_backup():
    """Exportar todos los pedidos activos en memoria como JSON para backup"""
    backup = _generar_backup_json()
    backup_json = json.dumps(backup, ensure_ascii=False, indent=2, default=str)
    output = io.BytesIO(backup_json.encode('utf-8'))
    output.seek(0)

    filename = f"backup_scanner_{now_mx_str('%Y%m%d_%H%M%S')}.json"
    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype='application/json')

@app.route("/download_backup_turso")
def download_backup_turso():
    """Descargar el ultimo backup automatico guardado en Turso"""
    try:
        backup = database.get_latest_backup_db()
        if not backup or not backup.get('backup_json'):
            return jsonify({"error": "No hay backups guardados en Turso"}), 404

        output = io.BytesIO(backup['backup_json'].encode('utf-8'))
        output.seek(0)
        fecha = backup.get('fecha', 'desconocida').replace('/', '').replace(' ', '_').replace(':', '')
        filename = f"backup_turso_{fecha}.json"
        return send_file(output, download_name=filename, as_attachment=True,
                         mimetype='application/json')
    except Exception as e:
        return jsonify({"error": f"Error: {str(e)}"}), 500

@app.route("/force_backup")
def force_backup():
    """Forzar guardado de backup en Turso ahora"""
    _guardar_backup_turso()
    backup = database.get_latest_backup_db()
    return jsonify({
        "success": True,
        "mensaje": "Backup guardado en Turso",
        "fecha": backup.get('fecha', '') if backup else '',
        "total_pedidos": backup.get('total_pedidos', 0) if backup else 0
    })

@app.route("/get_backup_info")
def get_backup_info():
    """Obtener info del ultimo backup guardado en Turso"""
    try:
        backup = database.get_latest_backup_db()
        if not backup:
            return jsonify({"hay_backup": False})
        return jsonify({
            "hay_backup": True,
            "fecha": backup.get('fecha', ''),
            "total_pedidos": backup.get('total_pedidos', 0),
            "created_at": backup.get('created_at', '')
        })
    except Exception as e:
        return jsonify({"hay_backup": False, "error": str(e)})

@app.route("/restore_backup", methods=["POST"])
def restore_backup():
    """Importar backup JSON y restaurar todos los pedidos activos"""
    global PEDIDOS_DB

    backup_file = request.files.get('backup_file')
    if not backup_file:
        return jsonify({"error": "Se requiere archivo de backup"}), 400

    try:
        content = backup_file.read().decode('utf-8')
        backup = json.loads(content)
    except Exception as e:
        return jsonify({"error": f"Error leyendo JSON: {str(e)}"}), 400

    if 'pedidos' not in backup:
        return jsonify({"error": "Formato de backup invalido"}), 400

    restaurados = 0
    errores = []

    for pid, ped_data in backup['pedidos'].items():
        try:
            pedido_cache = ped_data.get('pedido_cache', {})
            escaneos_raw = ped_data.get('escaneos_cache', {})
            info = ped_data.get('info', {})

            # Reconstruir escaneos_cache como defaultdict
            escaneos_cache = defaultdict(lambda: {'cantidad': 0, 'timestamp': [], 'scans': []})
            for lote, data in escaneos_raw.items():
                escaneos_cache[lote] = {
                    'cantidad': data.get('cantidad', 0),
                    'timestamp': data.get('timestamp', []),
                    'scans': data.get('scans', [])
                }

            PEDIDOS_DB[pid] = {
                'pedido_cache': pedido_cache,
                'escaneos_cache': escaneos_cache,
                'ultimos_scans': ped_data.get('ultimos_scans', {}),
                'modo_tarima': ped_data.get('modo_tarima', False),
                'tarima_pendiente': ped_data.get('tarima_pendiente', {}),
                'info': info
            }

            # Persistir en Turso/SQLite
            try:
                database.save_pedido_activo(pid, info, pedido_cache)
            except Exception as e:
                print(f"[DB] Error persistiendo pedido restaurado {pid}: {e}")

            restaurados += 1
        except Exception as e:
            errores.append(f"{pid}: {str(e)}")

    return jsonify({
        "success": True,
        "restaurados": restaurados,
        "errores": errores,
        "mensaje": f"{restaurados} pedidos restaurados correctamente"
    })

# ============================================================
# RUTAS: OPERADOR (CELULAR) - ESCANEAR QR
# ============================================================

@app.route("/scan/<pedido_id>")
def scan_page(pedido_id):
    if pedido_id not in PEDIDOS_DB:
        return "Pedido no encontrado", 404
    return render_template("operador_v3.html", pedido_id=pedido_id)

@app.route("/scan_qr/<pedido_id>", methods=["POST"])
def scan_qr(pedido_id):
    if pedido_id not in PEDIDOS_DB:
        return jsonify({"error": "Pedido no encontrado"}), 404

    db = PEDIDOS_DB[pedido_id]
    pedido_cache = db['pedido_cache']
    escaneos_cache = db['escaneos_cache']

    data = request.get_json()
    qr_data = data.get('qr_data', '')
    cantidad = data.get('cantidad', 1)
    usuario = data.get('usuario', '')

    if not qr_data:
        return jsonify({"error": "QR vacio"}), 400

    # Extraer lote del QR
    lote = None
    match = re.search(r'\(10\)(\d{10})', qr_data)
    if match:
        lote = match.group(1)

    if not lote:
        match = re.search(r'(?:LOTE|lote|Lote)[^\d]*(\d{6,12})', qr_data)
        if match:
            lote = match.group(1)

    if not lote:
        digits = re.findall(r'\d+', qr_data)
        if digits:
            for d in reversed(digits):
                if len(d) >= 6:
                    lote = d
                    break

    if not lote:
        return jsonify({"success": False, "error": "No se pudo extraer lote del QR", "qr": qr_data})

    timestamp = now_mx_str("%H:%M:%S")

    # ============================================================
    # OPERACION ATOMICA CON LOCK (multioperador safe)
    # Solo 1 operador puede escribir a la vez, pero la operacion
    # es rapida (< 1ms) asi que no hay delay percibido
    # ============================================================
    with SCAN_LOCK:
        # --- PREVENCION DE DOBLE SCAN ACCIDENTAL ---
        # Solo bloquear si el MISMO operador escanea el MISMO QR
        # en menos de 3 segundos (double-tap accidental del mismo operador)
        # Permitir QRs repetidos en cajas diferentes (normal en VPC)
        if 'ultimos_scans' not in db:
            db['ultimos_scans'] = {}  # {qr_key: {usuario: timestamp_epoch}}
        qr_key = qr_data.strip()[:200]
        ahora_epoch = now_mx().timestamp()
        if qr_key in db['ultimos_scans']:
            ultimo = db['ultimos_scans'][qr_key]
            ultimo_usuario = ultimo.get('usuario', '')
            ultimo_tiempo = ultimo.get('tiempo', 0)
            # Solo bloquear si mismo operador y menos de 3 segundos
            if ultimo_usuario == usuario and (ahora_epoch - ultimo_tiempo) < 3:
                return jsonify({
                    "success": False,
                    "duplicado": True,
                    "error": "QR YA ESCANEADO",
                    "lote": lote,
                    "descripcion": pedido_cache.get(lote, {}).get('descripcion', 'N/A')
                })
        db['ultimos_scans'][qr_key] = {'usuario': usuario, 'tiempo': ahora_epoch}

        # Iniciar timer automáticamente en primer scan (solo una vez)
        if db['info']['inicio_scan'] is None:
            db['info']['inicio_scan'] = now_mx_str("%d/%m/%Y %H:%M:%S")

        # Registrar operador (sin sobrescribir, acumular lista)
        if usuario:
            operadores_actuales = db['info'].get('usuarios_operadores', [])
            if usuario not in operadores_actuales:
                operadores_actuales.append(usuario)
                db['info']['usuarios_operadores'] = operadores_actuales
            # Mantener compatibilidad hacia atras (primer operador)
            if not db['info'].get('usuario_operador'):
                db['info']['usuario_operador'] = usuario

        # Guardar scan con usuario
        escaneos_cache[lote]['timestamp'].append(timestamp)
        escaneos_cache[lote]['scans'].append({
            'hora': timestamp,
            'qr': qr_data[:50],
            'cantidad': cantidad,
            'usuario': usuario
        })

        # --- MODO TARIMA (COSTCO / WALMART) ---
        # En modo tarima, las 3 primeras cajas son verificación (no se cuentan)
        # Después de 3 scans del mismo lote, se solicita el número de cajas por tarima
        if db.get('modo_tarima', False):
            if 'tarima_pendiente' not in db:
                db['tarima_pendiente'] = {}
            tarima_pend = db['tarima_pendiente']
            tarima_pend[lote] = tarima_pend.get(lote, 0) + 1

            if tarima_pend[lote] < 3:
                # Scans de verificación (1 y 2) - NO sumar al total
                # Marcar como verificación
                escaneos_cache[lote]['scans'][-1]['verificacion'] = True
                if lote in pedido_cache:
                    pedido_cache[lote]['escaneado'] = escaneos_cache[lote]['cantidad']

                return jsonify({
                    "success": True,
                    "verificacion_tarima": True,
                    "scan_verificacion": tarima_pend[lote],
                    "lote": lote,
                    "descripcion": pedido_cache.get(lote, {}).get('descripcion', 'N/A'),
                    "pedido": pedido_cache.get(lote, {}).get('pedido', 0),
                    "escaneado": escaneos_cache[lote]['cantidad'],
                    "timestamp": timestamp,
                    "mensaje": f"Verificación {tarima_pend[lote]}/3 - Escanea otra caja del mismo lote"
                })

            if tarima_pend[lote] >= 3:
                # 3 scans completados - solicitar cajas por tarima al operador
                # Reset contador para próxima tarima del mismo lote
                tarima_pend[lote] = 0
                # Marcar el 3er scan como verificación también
                escaneos_cache[lote]['scans'][-1]['verificacion'] = True
                if lote in pedido_cache:
                    pedido_cache[lote]['escaneado'] = escaneos_cache[lote]['cantidad']

                return jsonify({
                    "success": True,
                    "solicitar_tarima": True,
                    "lote": lote,
                    "descripcion": pedido_cache.get(lote, {}).get('descripcion', 'N/A'),
                    "pedido": pedido_cache.get(lote, {}).get('pedido', 0),
                    "escaneado": escaneos_cache[lote]['cantidad'],
                    "timestamp": timestamp,
                    "mensaje": "3 cajas verificadas. Ingresa cajas por tarima."
                })
        else:
            # Modo normal: sumar cada scan al total
            escaneos_cache[lote]['cantidad'] += cantidad
            if lote in pedido_cache:
                pedido_cache[lote]['escaneado'] = escaneos_cache[lote]['cantidad']

    # Persistir scan en base de datos (tiempo real)
    try:
        database.save_scan(pedido_id, lote, qr_data, cantidad, usuario, timestamp)
    except Exception as e:
        print(f"[DB] Error guardando scan: {e}")

    # Calcular estado DESPUES del lock (lectura es segura)
    if lote in pedido_cache:
        pedido = pedido_cache[lote]['pedido']
        escaneado = pedido_cache[lote]['escaneado']
        estado, color, texto = calcular_estado(pedido, escaneado)

        return jsonify({
            "success": True,
            "lote": lote,
            "descripcion": pedido_cache[lote]['descripcion'],
            "pedido": pedido,
            "escaneado": escaneado,
            "diferencia": escaneado - pedido,
            "estado": estado,
            "color": color,
            "texto_estado": texto,
            "timestamp": timestamp,
            "es_nuevo": len(escaneos_cache[lote]['scans']) == 1,
            "total_scans": len(escaneos_cache[lote]['scans'])
        })
    else:
        inv_data = INVENTARIO_DB.get(lote, {})
        return jsonify({
            "success": True,
            "lote": lote,
            "descripcion": inv_data.get('descripcion', 'NO EN PEDIDO'),
            "pedido": 0,
            "escaneado": escaneos_cache[lote]['cantidad'],
            "diferencia": escaneos_cache[lote]['cantidad'],
            "estado": "SOBRANTE_INESPERADO",
            "color": "amarillo",
            "texto_estado": f"SOBRANTE INESPERADO",
            "timestamp": timestamp,
            "es_nuevo": True,
            "alerta": "Este lote NO esta en el picking!",
            "total_scans": len(escaneos_cache[lote]['scans'])
        })

@app.route("/registrar_tarima/<pedido_id>", methods=["POST"])
def registrar_tarima(pedido_id):
    """Registrar cajas por tarima en modo tarima (COSTCO/WALMART)"""
    if pedido_id not in PEDIDOS_DB:
        return jsonify({"error": "Pedido no encontrado"}), 404

    db = PEDIDOS_DB[pedido_id]
    if not db.get('modo_tarima', False):
        return jsonify({"error": "Este pedido no esta en modo tarima"}), 400

    data = request.get_json()
    lote = data.get('lote', '')
    cajas_por_tarima = data.get('cajas_por_tarima', 0)
    usuario = data.get('usuario', '')

    if not lote or cajas_por_tarima <= 0:
        return jsonify({"error": "Lote y cajas_por_tarima son requeridos"}), 400

    pedido_cache = db['pedido_cache']
    escaneos_cache = db['escaneos_cache']
    timestamp = now_mx_str("%H:%M:%S")

    with SCAN_LOCK:
        # Sumar las cajas por tarima al total
        escaneos_cache[lote]['cantidad'] += cajas_por_tarima
        escaneos_cache[lote]['timestamp'].append(timestamp)
        escaneos_cache[lote]['scans'].append({
            'hora': timestamp,
            'qr': f'TARIMA:{cajas_por_tarima}',
            'cantidad': cajas_por_tarima,
            'usuario': usuario,
            'es_tarima': True
        })

        if lote in pedido_cache:
            pedido_cache[lote]['escaneado'] = escaneos_cache[lote]['cantidad']

    # Persistir
    try:
        database.save_scan(pedido_id, lote, f'TARIMA:{cajas_por_tarima}', cajas_por_tarima, usuario, timestamp)
    except Exception as e:
        print(f"[DB] Error guardando tarima: {e}")

    if lote in pedido_cache:
        pedido = pedido_cache[lote]['pedido']
        escaneado = pedido_cache[lote]['escaneado']
        estado, color, texto = calcular_estado(pedido, escaneado)

        return jsonify({
            "success": True,
            "lote": lote,
            "descripcion": pedido_cache[lote]['descripcion'],
            "pedido": pedido,
            "escaneado": escaneado,
            "diferencia": escaneado - pedido,
            "estado": estado,
            "color": color,
            "texto_estado": texto,
            "timestamp": timestamp,
            "cajas_registradas": cajas_por_tarima,
            "total_scans": len(escaneos_cache[lote]['scans'])
        })
    else:
        return jsonify({
            "success": True,
            "lote": lote,
            "descripcion": "NO EN PEDIDO",
            "escaneado": escaneos_cache[lote]['cantidad'],
            "cajas_registradas": cajas_por_tarima,
            "timestamp": timestamp
        })

@app.route("/finalizar/<pedido_id>")
def finalizar(pedido_id):
    if pedido_id not in PEDIDOS_DB:
        return jsonify({"error": "Pedido no encontrado"}), 404

    db = PEDIDOS_DB[pedido_id]
    pedido_cache = db['pedido_cache']
    escaneos_cache = db['escaneos_cache']

    # Auto-finalizar si no está finalizado
    try:
        if db['info']['fin_scan'] is None:
            fin = now_mx()
            db['info']['fin_scan'] = fin.strftime("%d/%m/%Y %H:%M:%S")

            inicio_str = db['info']['inicio_scan']
            if inicio_str:
                # Parse inicio como naive y luego darle timezone de MX para comparar correctamente
                inicio_naive = datetime.strptime(inicio_str, "%d/%m/%Y %H:%M:%S")
                inicio = inicio_naive.replace(tzinfo=MX_TZ)
                diff = fin - inicio
                if diff.total_seconds() < 0:
                    diff = abs(diff)
                horas = int(diff.total_seconds() // 3600)
                minutos = int((diff.total_seconds() % 3600) // 60)
                segundos = int(diff.total_seconds() % 60)
                db['info']['tiempo_total'] = f"{horas:02d}:{minutos:02d}:{segundos:02d}"

        # Si tiempo_total sigue vacio pero tenemos inicio y fin, recalcular (reparar pedidos afectados)
        if not db['info']['tiempo_total'] and db['info']['inicio_scan'] and db['info']['fin_scan']:
            inicio_naive = datetime.strptime(db['info']['inicio_scan'], "%d/%m/%Y %H:%M:%S")
            fin_naive = datetime.strptime(db['info']['fin_scan'], "%d/%m/%Y %H:%M:%S")
            diff = fin_naive - inicio_naive
            if diff.total_seconds() < 0:
                diff = abs(diff)
            horas = int(diff.total_seconds() // 3600)
            minutos = int((diff.total_seconds() % 3600) // 60)
            segundos = int(diff.total_seconds() % 60)
            db['info']['tiempo_total'] = f"{horas:02d}:{minutos:02d}:{segundos:02d}"

        # Guardar en estadísticas (solo si no se habia guardado antes)
        ya_en_historial = any(h.get('pedido_id') == pedido_id for h in USO_STATS['historial'])
        if not ya_en_historial:
            total_escaneado = sum(v['escaneado'] for v in pedido_cache.values())
            USO_STATS['cajas_escaneadas_totales'] += total_escaneado
            USO_STATS['historial'].append({
                'pedido_id': pedido_id,
                'cliente': db['info']['header_data'].get('cliente', ''),
                'destinatario': db['info']['header_data'].get('destinatario', ''),
                'entrega_nro': db['info']['header_data'].get('entrega_nro', ''),
                'usuario': db['info']['usuario_operador'],
                'inicio': db['info']['inicio_scan'],
                'fin': db['info']['fin_scan'],
                'tiempo_total': db['info']['tiempo_total'],
                'cajas_escaneadas': total_escaneado
            })
    except Exception as e:
        print(f"[ERROR] finalizar calcular tiempo: {e}")
        db['info']['tiempo_total'] = db['info']['tiempo_total'] or '--'

    resultados = []
    ok = faltantes = sobrantes = 0

    for lote, data in pedido_cache.items():
        estado, color, texto = calcular_estado(data['pedido'], data['escaneado'])
        if color == 'verde': ok += 1
        elif color == 'rojo': faltantes += 1
        elif color == 'amarillo': sobrantes += 1

        resultados.append({
            'codigo': data['codigo'],
            'descripcion': data['descripcion'],
            'lote': lote,
            'calibre': data['calibre'],
            'categoria': data['categoria'],
            'pedido': data['pedido'],
            'escaneado': data['escaneado'],
            'diferencia': data['escaneado'] - data['pedido'],
            'estado': texto,
            'color': color
        })

    sobrantes_pdf = 0
    for lote, data in escaneos_cache.items():
        if lote not in pedido_cache:
            sobrantes_pdf += 1
            inv_data = INVENTARIO_DB.get(lote, {})
            resultados.append({
                'codigo': inv_data.get('codigo', 'NO EN PEDIDO'),
                'descripcion': inv_data.get('descripcion', 'Producto no listado'),
                'lote': lote,
                'calibre': inv_data.get('calibre', '-'),
                'categoria': inv_data.get('categoria', '-'),
                'pedido': 0,
                'escaneado': data['cantidad'],
                'diferencia': data['cantidad'],
                'estado': 'SOBRANTE INESPERADO',
                'color': 'amarillo'
            })

    def sort_key(r):
        if r['color'] == 'rojo': return 0
        if r['color'] == 'amarillo': return 1
        return 2
    resultados.sort(key=sort_key)

    # ============================================================
    # METRICAS - MISMA LOGICA QUE get_status()
    # ============================================================

    # Separar articulos del pedido vs fuera de pedido
    articulos_del_pedido = [r for r in resultados if r['estado'] != 'SOBRANTE INESPERADO']

    # Articulos (solo del pedido)
    articulos_ok = sum(1 for r in articulos_del_pedido if r['escaneado'] > 0)
    articulos_faltan = sum(1 for r in articulos_del_pedido if r['escaneado'] == 0)
    total_articulos_pedido = len(articulos_del_pedido)

    # Cajas
    cajas_ok = sum(r['escaneado'] for r in articulos_del_pedido)
    cajas_escaneadas = sum(r['escaneado'] for r in resultados)
    cajas_fuera_pedido = sum(r['escaneado'] for r in resultados if r['estado'] == 'SOBRANTE INESPERADO')
    articulos_fuera_pedido = sum(1 for r in resultados if r['estado'] == 'SOBRANTE INESPERADO')

    # Nueva metrica: Cajas Faltantes por escanear
    cajas_faltantes_por_escanear = max(0, sum(r['pedido'] for r in articulos_del_pedido) - cajas_ok)

    # GUARDAR EN BASE DE DATOS (persistencia del historial)
    resumen_db = {
        'total_articulos': total_articulos_pedido,
        'total_pedido': sum(r['pedido'] for r in articulos_del_pedido),
        'total_escaneado': cajas_escaneadas,
        'cajas_ok': cajas_ok,
        'cajas_fuera_pedido': cajas_fuera_pedido,
        'ok': articulos_ok,
        'faltantes': articulos_faltan
    }
    try:
        database.finalizar_pedido_db(pedido_id, db['info'], db['pedido_cache'], db['escaneos_cache'], resultados, resumen_db)
    except Exception as e:
        print(f"[DB] Error finalizando pedido: {e}")

    return jsonify({
        "success": True,
        "pedido_id": pedido_id,
        "resumen": {
            "total_articulos": total_articulos_pedido,
            "total_pedido": sum(r['pedido'] for r in articulos_del_pedido),
            "total_escaneado": cajas_escaneadas,
            "cajas_ok": cajas_ok,
            "cajas_escaneadas": cajas_escaneadas,
            "cajas_fuera_pedido": cajas_fuera_pedido,
            "cajas_faltantes_por_escanear": cajas_faltantes_por_escanear,
            "articulos_fuera_pedido": articulos_fuera_pedido,
            "ok": articulos_ok,
            "faltantes": articulos_faltan,
            "sobrantes": 0,
            "inicio_scan": db['info']['inicio_scan'],
            "fin_scan": db['info']['fin_scan'],
            "tiempo_total": db['info']['tiempo_total'],
            "usuario": db['info']['usuario_operador'],
            "header_data": db['info']['header_data']
        },
        "resultados": resultados
    })

# ============================================================
# DESCARGAS: EXCEL Y PDF MANIFIESTO CON DATOS DEL HEADER
# ============================================================

@app.route("/download/excel/<pedido_id>")
def download_excel(pedido_id):
    if pedido_id not in PEDIDOS_DB:
        return "Pedido no encontrado", 404

    db = PEDIDOS_DB[pedido_id]
    pedido_cache = db['pedido_cache']
    escaneos_cache = db['escaneos_cache']
    info = db['info']
    header = info['header_data']

    resultados = []
    for lote, data in pedido_cache.items():
        estado, color, texto = calcular_estado(data['pedido'], data['escaneado'])
        resultados.append({
            'Codigo': data['codigo'],
            'Descripcion': data['descripcion'],
            'Lote': lote,
            'Calibre': data['calibre'],
            'Categoria': data['categoria'],
            'Cantidad_Pedido': data['pedido'],
            'Cantidad_Escaneada': data['escaneado'],
            'Diferencia': data['escaneado'] - data['pedido'],
            'Estado': texto
        })

    for lote, data in escaneos_cache.items():
        if lote not in pedido_cache:
            inv_data = INVENTARIO_DB.get(lote, {})
            resultados.append({
                'Codigo': inv_data.get('codigo', 'NO EN PEDIDO'),
                'Descripcion': inv_data.get('descripcion', 'Producto no listado'),
                'Lote': lote,
                'Calibre': inv_data.get('calibre', '-'),
                'Categoria': inv_data.get('categoria', '-'),
                'Cantidad_Pedido': 0,
                'Cantidad_Escaneada': data['cantidad'],
                'Diferencia': data['cantidad'],
                'Estado': 'SOBRANTE INESPERADO'
            })

    # Separar resultados en dos grupos
    items_en_pedido = [r for r in resultados if r['Estado'] != 'SOBRANTE INESPERADO']
    items_fuera_pedido = [r for r in resultados if r['Estado'] == 'SOBRANTE INESPERADO']

    # Ordenar: Faltan primero
    def estado_sort(e):
        if 'FALTAN' in e: return 0
        if 'SOBRANTE' in e or 'SOBRAN' in e: return 1
        return 2
    resultados.sort(key=lambda x: estado_sort(x['Estado']))

    wb = Workbook()

    # Hoja 1: Resumen Embarque
    ws1 = wb.active
    ws1.title = 'Resumen Embarque'

    ws1.cell(row=1, column=1, value='CLIENTE')
    ws1.cell(row=1, column=2, value=header.get('cliente', ''))
    ws1.cell(row=2, column=1, value='DESTINATARIO')
    ws1.cell(row=2, column=2, value=header.get('destinatario', ''))
    ws1.cell(row=3, column=1, value='ENTREGA NRO')
    ws1.cell(row=3, column=2, value=header.get('entrega_nro', ''))
    ws1.cell(row=4, column=1, value='FECHA/HORA CITA')
    ws1.cell(row=4, column=2, value=header.get('fecha_hora_cita', ''))
    ws1.cell(row=5, column=1, value='FECHA/HORA IMPRESION')
    ws1.cell(row=5, column=2, value=header.get('fecha_hora_impresion', ''))
    ws1.cell(row=6, column=1, value='')
    ws1.cell(row=7, column=1, value='')

    headers = ['Codigo', 'Descripcion', 'Lote', 'Calibre', 'Categoria',
               'Cantidad_Pedido', 'Cantidad_Escaneada', 'Diferencia', 'Estado']
    ws1.append(headers)

    for r in resultados:
        ws1.append([
            r['Codigo'], r['Descripcion'], r['Lote'], r['Calibre'], r['Categoria'],
            r['Cantidad_Pedido'], r['Cantidad_Escaneada'], r['Diferencia'], r['Estado']
        ])

    # Hoja 2: Resumen General
    ws2 = wb.create_sheet('Resumen General')

    ok = sum(1 for r in resultados if 'OK' in r['Estado'])
    faltan = sum(1 for r in resultados if 'FALTAN' in r['Estado'])
    sobran = sum(1 for r in resultados if 'SOBRAN' in r['Estado'] or 'SOBRANTE' in r['Estado'])
    total_pedido = sum(r['Cantidad_Pedido'] for r in resultados)
    total_escan = sum(r['Cantidad_Escaneada'] for r in resultados)

    ws2.cell(row=1, column=1, value='CLIENTE')
    ws2.cell(row=1, column=2, value=header.get('cliente', ''))
    ws2.cell(row=2, column=1, value='DESTINATARIO')
    ws2.cell(row=2, column=2, value=header.get('destinatario', ''))
    ws2.cell(row=3, column=1, value='ENTREGA NRO')
    ws2.cell(row=3, column=2, value=header.get('entrega_nro', ''))
    ws2.cell(row=4, column=1, value='')
    ws2.cell(row=5, column=1, value='')
    ws2.append(['Metrica', 'Valor'])
    ws2.append(['Pedido ID', pedido_id])
    ws2.append(['Total Articulos', len(resultados)])
    ws2.append(['Total Cajas Pedido', total_pedido])
    ws2.append(['Total Cajas Escaneadas', total_escan])
    ws2.append(['Correctos (OK)', ok])
    ws2.append(['Faltantes', faltan])
    ws2.append(['Sobrantes', sobran])
    ws2.append(['Inicio Scan', info.get('inicio_scan', '-')])
    ws2.append(['Fin Scan', info.get('fin_scan', '-')])
    ws2.append(['Tiempo Total', info.get('tiempo_total', '-')])
    ws2.append(['Usuario Operador', info.get('usuario_operador', '-')])
    ws2.append(['Fecha', now_mx_str("%d/%m/%Y %H:%M")])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=f"Resumen_Embarque_{pedido_id}.xlsx", as_attachment=True)

@app.route("/download/pdf/<pedido_id>")
def download_pdf(pedido_id):
    if pedido_id not in PEDIDOS_DB:
        return "Pedido no encontrado", 404

    db = PEDIDOS_DB[pedido_id]
    pedido_cache = db['pedido_cache']
    escaneos_cache = db['escaneos_cache']
    info = db['info']
    header = info['header_data']

    resultados = []
    for lote, data in pedido_cache.items():
        estado, color, texto = calcular_estado(data['pedido'], data['escaneado'])
        resultados.append({
            'codigo': data['codigo'],
            'descripcion': data['descripcion'],
            'lote': lote,
            'calibre': data['calibre'],
            'categoria': data['categoria'],
            'pedido': data['pedido'],
            'escaneado': data['escaneado'],
            'diferencia': data['escaneado'] - data['pedido'],
            'estado': texto,
            'color': color
        })

    sobrantes_pdf = 0
    for lote, data in escaneos_cache.items():
        if lote not in pedido_cache:
            sobrantes_pdf += 1
            inv_data = INVENTARIO_DB.get(lote, {})
            resultados.append({
                'codigo': inv_data.get('codigo', 'NO EN PEDIDO'),
                'descripcion': inv_data.get('descripcion', 'Producto no listado'),
                'lote': lote,
                'calibre': inv_data.get('calibre', '-'),
                'categoria': inv_data.get('categoria', '-'),
                'pedido': 0,
                'escaneado': data['cantidad'],
                'diferencia': data['cantidad'],
                'estado': 'SOBRANTE INESPERADO',
                'color': 'amarillo'
            })

    def sort_key(r):
        if r['color'] == 'rojo': return 0
        if r['color'] == 'amarillo': return 1
        return 2
    resultados.sort(key=sort_key)

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter),
                            rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=20,
                                 textColor=colors.HexColor('#1a5276'), spaceAfter=12, alignment=1)
    subtitle_style = ParagraphStyle('CustomSub', parent=styles['Normal'], fontSize=12,
                                  textColor=colors.HexColor('#666'), alignment=1, spaceAfter=20)

    elements = []

    # LOGO
    try:
        logo_path = os.path.join(app.root_path, 'static', 'logo-vpc.png')
        if os.path.exists(logo_path):
            img = Image(logo_path, width=2*inch, height=1*inch)
            elements.append(img)
    except:
        pass

    elements.append(Paragraph("<b>MANIFIESTO DE REVISION DE EMBARQUE</b>", title_style))

    # Header data
    header_info = f"""
    <b>Cliente:</b> {header.get('cliente', 'N/A')} | 
    <b>Destinatario:</b> {header.get('destinatario', 'N/A')} | 
    <b>Entrega Nro:</b> {header.get('entrega_nro', 'N/A')}<br>
    <b>Cita:</b> {header.get('fecha_hora_cita', 'N/A')} | 
    <b>Impresión:</b> {header.get('fecha_hora_impresion', 'N/A')}
    """
    elements.append(Paragraph(header_info, subtitle_style))

    tiempo_info = ""
    if info.get('tiempo_total'):
        tiempo_info = f" | <b>Tiempo:</b> {info['tiempo_total']}"

    elements.append(Paragraph(
        f"<b>Pedido ID:</b> {pedido_id} | <b>Fecha:</b> {now_mx_str('%d/%m/%Y %H:%M')}{tiempo_info}", 
        subtitle_style))
    elements.append(Spacer(1, 0.1*inch))

    ok = sum(1 for r in resultados if r['color'] == 'verde')
    rojo = sum(1 for r in resultados if r['color'] == 'rojo')
    amarillo = sum(1 for r in resultados if r['color'] == 'amarillo')
    sobrantes_pdf = amarillo  # Usar el conteo de la lista resultados
    total_pedido = sum(r['pedido'] for r in resultados)
    total_escan = sum(r['escaneado'] for r in resultados)

    resumen_data = [
        ['RESUMEN EJECUTIVO', '', '', '', '', ''],
        ['Total Articulos', 'Cajas Pedido', 'Cajas Escaneadas', '✅ OK', '🔴 Faltantes', '🟡 Sobrantes'],
        [str(len(resultados)), str(int(total_pedido)), str(int(total_escan)), str(ok), str(rojo), str(amarillo)]
    ]

    if info.get('inicio_scan'):
        resumen_data.append(['Inicio', 'Fin', 'Tiempo Total', 'Usuario', '', ''])
        resumen_data.append([
            info.get('inicio_scan', '-'), 
            info.get('fin_scan', '-'), 
            info.get('tiempo_total', '-'),
            info.get('usuario_operador', '-'),
            '', ''
        ])

    resumen_table = Table(resumen_data, colWidths=[1.5*inch]*6)
    resumen_style = [
        ('SPAN', (0,0), (-1,0)),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a5276')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (0,0), 14),
        ('FONTSIZE', (0,1), (-1,1), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 10),
        ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#2980b9')),
        ('TEXTCOLOR', (0,1), (-1,1), colors.whitesmoke),
        ('BACKGROUND', (0,2), (-1,2), colors.HexColor('#d6eaf8')),
        ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#1a5276')),
        ('FONTNAME', (0,2), (-1,2), 'Helvetica-Bold'),
        ('FONTSIZE', (0,2), (-1,2), 12),
    ]

    if info.get('inicio_scan'):
        resumen_style.append(('BACKGROUND', (0,3), (-1,3), colors.HexColor('#2980b9')))
        resumen_style.append(('TEXTCOLOR', (0,3), (-1,3), colors.whitesmoke))
        resumen_style.append(('BACKGROUND', (0,4), (-1,4), colors.HexColor('#d6eaf8')))
        resumen_style.append(('FONTNAME', (0,4), (-1,4), 'Helvetica-Bold'))

    resumen_table.setStyle(TableStyle(resumen_style))
    elements.append(resumen_table)
    elements.append(Spacer(1, 0.3*inch))

    elements.append(Paragraph("<b>DETALLE POR ARTICULO</b>", styles['Heading2']))
    elements.append(Spacer(1, 0.1*inch))

    table_data = [['Codigo', 'Descripcion', 'Lote', 'Cal.', 'Cat.', 'Pedido', 'Escaneado', 'Dif.', 'Estado']]
    for r in resultados:
        table_data.append([
            r['codigo'], r['descripcion'][:40], r['lote'], r['calibre'], r['categoria'],
            str(int(r['pedido'])), str(int(r['escaneado'])), str(int(r['diferencia'])), r['estado']
        ])

    col_widths = [1.1*inch, 2.4*inch, 0.9*inch, 0.5*inch, 0.5*inch, 0.7*inch, 0.8*inch, 0.5*inch, 1.2*inch]
    table = Table(table_data, colWidths=col_widths, repeatRows=2)

    style_commands = [
        ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#1a5276')),
        ('TEXTCOLOR', (0,1), (-1,1), colors.whitesmoke),
        ('ALIGN', (0,1), (-1,-1), 'CENTER'),
        ('ALIGN', (1,2), (1,-1), 'LEFT'),
        ('FONTNAME', (0,1), (-1,1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,1), (-1,1), 9),
        ('BOTTOMPADDING', (0,1), (-1,1), 8),
        ('GRID', (0,1), (-1,-1), 0.5, colors.grey),
        ('FONTNAME', (0,2), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,2), (-1,-1), 8),
        ('VALIGN', (0,1), (-1,-1), 'MIDDLE'),
    ]

    for i, r in enumerate(resultados, start=2):
        if r['color'] == 'rojo':
            style_commands.append(('BACKGROUND', (0,i), (-1,i), colors.HexColor('#fadbd8')))
        elif r['color'] == 'amarillo':
            style_commands.append(('BACKGROUND', (0,i), (-1,i), colors.HexColor('#fcf3cf')))
        elif r['color'] == 'verde':
            style_commands.append(('BACKGROUND', (0,i), (-1,i), colors.HexColor('#d5f5e3')))

    table.setStyle(TableStyle(style_commands))
    elements.append(table)

    elements.append(Spacer(1, 0.3*inch))
    elements.append(Paragraph("<b>Observaciones:</b> _________________________________________________________________", styles['Normal']))
    elements.append(Spacer(1, 0.1*inch))
    elements.append(Paragraph("<b>Reviso:</b> _________________________    <b>Firma:</b> _________________________", styles['Normal']))

    # Leyenda
    elements.append(Spacer(1, 0.2*inch))
    elements.append(Paragraph("<i>Desarrollado por Luis Hori</i>", ParagraphStyle('Footer', parent=styles['Normal'], 
                                                                                fontSize=8, textColor=colors.grey, alignment=1)))

    doc.build(elements)
    output.seek(0)

    return send_file(output, download_name=f"Manifiesto_Embarque_{pedido_id}.pdf", as_attachment=True)


# ============================================================
# RUTA: GUARDAR EN GOOGLE DRIVE
# ============================================================

@app.route("/guardar_drive/<pedido_id>", methods=["POST"])
def guardar_drive(pedido_id):
    if pedido_id not in PEDIDOS_DB:
        return jsonify({"error": "Pedido no encontrado"}), 404

    db = PEDIDOS_DB[pedido_id]
    header = db['info']['header_data']

    # Verificar si Drive está configurado
    service = get_drive_service()
    if not service:
        return jsonify({"error": "Google Drive no configurado. Verifica GOOGLE_DRIVE_CREDENTIALS en Environment"}), 500

    data = request.get_json() or {}
    tipo = data.get('tipo', 'pdf')

    try:
        temp_path = None
        file_name = None
        mime_type = None

        if tipo == 'pdf':
            # Generar PDF directamente
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors

            temp_path = os.path.join(UPLOAD_FOLDER, f"temp_drive_{pedido_id}.pdf")
            file_name = f"Manifiesto_{pedido_id}.pdf"
            mime_type = "application/pdf"

            # Generar PDF (código simplificado - reutilizar lógica existente)
            # Por ahora, usamos el archivo ya generado si existe
            # TODO: Generar PDF inline

            # Fallback: generar PDF básico
            doc = SimpleDocTemplate(temp_path, pagesize=landscape(letter))
            elements = []
            styles = getSampleStyleSheet()
            elements.append(Paragraph(f"<b>Manifiesto {pedido_id}</b>", styles['Heading1']))
            elements.append(Paragraph(f"Cliente: {header.get('cliente', 'N/A')}", styles['Normal']))
            elements.append(Paragraph(f"Destinatario: {header.get('destinatario', 'N/A')}", styles['Normal']))
            doc.build(elements)

        elif tipo == 'excel':
            # Generar Excel directamente
            from openpyxl import Workbook

            temp_path = os.path.join(UPLOAD_FOLDER, f"temp_drive_{pedido_id}.xlsx")
            file_name = f"Resumen_{pedido_id}.xlsx"
            mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

            wb = Workbook()
            ws = wb.active
            ws.title = 'Resumen'
            ws.append(['CLIENTE', header.get('cliente', '')])
            ws.append(['DESTINATARIO', header.get('destinatario', '')])
            ws.append(['ENTREGA NRO', header.get('entrega_nro', '')])
            wb.save(temp_path)

        else:
            return jsonify({"error": "Tipo no válido. Use 'pdf' o 'excel'"}), 400

        # Subir a Drive
        result = guardar_manifiesto_en_drive(pedido_id, header, temp_path, file_name, mime_type)

        # Limpiar temporal
        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)

        if result:
            return jsonify({"success": True, "link": result['link'], "mensaje": f"{tipo.upper()} guardado en Drive"})
        else:
            return jsonify({"error": "No se pudo guardar en Drive. Verifica permisos de la carpeta."}), 500

    except Exception as e:
        return jsonify({"error": f"Error: {str(e)}"}), 500


# ============================================================
# CONFIGURACIÓN PARA PRODUCCIÓN (Render)
# ============================================================
import os

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
else:
    port = int(os.environ.get("PORT", 5000))
    app.debug = False
